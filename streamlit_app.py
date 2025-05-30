# app.py
import streamlit as st
import pandas as pd
import time
import re
from io import BytesIO
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from concurrent.futures import ThreadPoolExecutor, as_completed

# → Clé OpenAI
client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

# — Page config
st.set_page_config(page_title="AI Excel Processor", layout="wide")

# — INITIALISATION session_state
for key, default in [
    ("model", "gpt-4o-mini"),
    ("temperature", 0.0),
    ("rate_limit", 1.0),
    ("stop_flag", False),
    ("error_rows", []),
    ("log_entries", []),
    ("last_processed", -1),
    ("completed", False),
    ("prompt_text", ""),
    ("cols_to_insert", []),
    ("templates", []),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# — Sidebar : Réglages
with st.sidebar:
    st.header("⚙️ Configurations")

    presets = {
        "Exploration rapide": ("gpt-3.5-turbo", 0.7, 0.2),
        "Production stable":    ("gpt-4o-mini", 0.0, 1.0),
    }
    st.selectbox("Préconfiguration", [""] + list(presets.keys()), key="preset")
    if st.button("🔄 Appliquer preset"):
        if st.session_state.preset in presets:
            m, t, r = presets[st.session_state.preset]
            st.session_state.model = m
            st.session_state.temperature = t
            st.session_state.rate_limit = r

    st.selectbox("Modèle", ["gpt-4o-mini", "gpt-3.5-turbo"], key="model")
    st.slider("Température", 0.0, 1.0, key="temperature")
    st.number_input("Pause entre requêtes (s)", 0.0, step=0.1, key="rate_limit")

    st.markdown("---")
    st.header("🚀 Exécution")
    run_full = st.button("▶️ Lancer le fichier complet")
    run_test = st.button("⚡ Test rapide (5 lignes)")
    if st.button("⏹️ Stop"):
        st.session_state.stop_flag = True

# — Cache persistant pour les prompts
@st.cache_resource
def _prompt_cache():
    return {}
prompt_cache = _prompt_cache()

# — Upload Excel & cache workbook
uploaded = st.file_uploader("📂 Chargez votre fichier Excel", type=["xlsx"])
if not uploaded:
    st.stop()

if (
    "bytes" not in st.session_state
    or st.session_state.filename != uploaded.name
    or st.session_state.filesize != uploaded.size
):
    st.session_state.bytes    = uploaded.read()
    st.session_state.filename = uploaded.name
    st.session_state.filesize = uploaded.size

    wb = load_workbook(
        filename=BytesIO(st.session_state.bytes),
        read_only=False,
        data_only=False
    )
    st.session_state.wb = wb
    st.session_state.sheet_names = wb.sheetnames

# — Choix de la feuille & lazy load
sheet = st.selectbox("🗂 Onglet à traiter", st.session_state.sheet_names)

@st.cache_data
def _load_sheet(bts, sht):
    return pd.read_excel(BytesIO(bts), engine="openpyxl", sheet_name=sht)

df = _load_sheet(st.session_state.bytes, sheet).copy()

# — Filtrage global
filter_kw = st.text_input("🔍 Filtrer (mot-clé)", "")
if filter_kw:
    df = df[df.apply(lambda row: row.astype(str).str.contains(filter_kw, case=False).any(), axis=1)]

# — Éditeur de données
st.markdown("### ✏️ Éditeur de données")
if hasattr(st, "data_editor"):
    df = st.data_editor(df, num_rows="dynamic")
else:
    df = st.experimental_data_editor(df, num_rows="dynamic")

st.markdown(f"**{sheet}** : {df.shape[0]} lignes × {df.shape[1]} colonnes")

# --- Prompt & placeholders ---
st.markdown("### 📝 Rédigez votre prompt")
st.text_area("Utilisez #Colonne# pour les placeholders", height=200, key="prompt_text")

st.markdown("### ➕ Colonnes à insérer")
st.multiselect("Votre choix", options=list(df.columns), key="cols_to_insert")

def _insert_all_ph():
    for c in st.session_state.cols_to_insert:
        ph = f"#{c}#"
        if ph not in st.session_state.prompt_text:
            st.session_state.prompt_text += ph + " "
st.button("➕ Ajouter tous les placeholders", on_click=_insert_all_ph)

# Validation placeholders
prompt_tpl   = st.session_state.prompt_text
placeholders = re.findall(r"#([^#]+)#", prompt_tpl)
invalid      = [c for c in placeholders if c not in df.columns]
if invalid:
    st.error(f"Colonnes invalides : {', '.join(invalid)}")
    st.stop()
if not placeholders:
    st.warning("Aucun placeholder détecté.")

# Aperçu sur la 1ʳᵉ ligne
if placeholders and not df.empty:
    st.markdown("#### 📄 Aperçu (1ʳᵉ ligne)")
    row0 = df.iloc[0].to_dict()
    filled0 = prompt_tpl
    for c in placeholders:
        filled0 = filled0.replace(f"#{c}#", str(row0.get(c, "")))
    st.text_area("Prompt exemple", filled0, height=100, disabled=True)

# Templates
st.markdown("### 🎁 Templates de prompt")
tname = st.text_input("Nom du template")
if st.button("💾 Sauvegarder template") and tname:
    st.session_state.templates.append({
        "name": tname,
        "prompt": prompt_tpl,
        "cols": st.session_state.cols_to_insert.copy()
    })
    st.success("Template sauvegardé.")

names = [t["name"] for t in st.session_state.templates]
sel   = st.selectbox("Charger un template", options=[""] + names)
if sel and st.button("📂 Charger template"):
    tmpl = next(t for t in st.session_state.templates if t["name"] == sel)
    st.session_state.prompt_text    = tmpl["prompt"]
    st.session_state.cols_to_insert = tmpl["cols"]
    st.experimental_rerun()

# — Colonne résultat
output_col = st.text_input("Nom de la colonne résultat", "Réponse IA")
if output_col not in df.columns:
    df[output_col] = ""

# — Préparer les variables pour la boucle
model       = st.session_state.model
temperature = st.session_state.temperature
rate_limit  = st.session_state.rate_limit

def call_chat(prompt: str) -> str:
    if prompt in prompt_cache:
        return prompt_cache[prompt]
    try:
        resp = client.chat.completions.create(
            model=model,
            temperature=temperature,
            messages=[
                {"role": "system", "content": "Vous êtes un assistant utile et précis."},
                {"role": "user",   "content": prompt}
            ]
        )
        text = resp.choices[0].message.content.strip()
    except Exception as e:
        text = f"Erreur API : {e}"
    prompt_cache[prompt] = text
    return text

def _process_row(i, row):
    data   = {c: ("" if pd.isna(v) else str(v)) for c, v in row.items()}
    filled = prompt_tpl
    for c in placeholders:
        filled = filled.replace(f"#{c}#", data.get(c, ""))
    start  = time.time()
    resp   = call_chat(filled)
    dur    = time.time() - start
    status = "error" if resp.startswith("Erreur API") else "success"
    return i, resp, status, dur, filled

# — Sélection des indices
to_process = list(df.index) if run_full else list(df.index[:5]) if run_test else []

if to_process:
    st.session_state.stop_flag = False
    error_rows_local = []
    log_entries_local = []
    total = len(to_process)
    prog  = st.progress(0)
    live  = st.empty()

    workers = max(1, int(1 / rate_limit)) if rate_limit > 0 else 1
    with ThreadPoolExecutor(max_workers=workers) as exe:
        futures = {exe.submit(_process_row, i, df.loc[i]): i for i in to_process}
        done = 0
        for fut in as_completed(futures):
            if st.session_state.stop_flag:
                st.warning("⚠️ Traitement interrompu.")
                break
            i, resp, status, dur, filled = fut.result()
            df.at[i, output_col] = resp
            log_entries_local.append({
                "index":   i,
                "prompt":  filled,
                "status":  status,
                "duration": dur
            })
            if status == "error":
                error_rows_local.append(i)
            done += 1
            prog.progress(done / total)
            live.dataframe(df.head(50), height=250)

    # Update session_state après la boucle
    st.session_state.error_rows    = error_rows_local
    st.session_state.log_entries   = log_entries_local
    st.session_state.last_processed = to_process[done - 1] if done else -1
    st.session_state.completed      = True
    st.success("✅ Traitement fini.")

# — Relance des erreurs
if st.session_state.error_rows:
    if st.button("🔄 Réessayer erreurs"):
        run_full = False
        run_test = False
        to_process = st.session_state.error_rows.copy()
        st.session_state.error_rows = []
        st.experimental_rerun()

# — Préparer le téléchargement
if st.session_state.completed:
    buf = BytesIO()
    wb  = st.session_state.wb
    if sheet in wb.sheetnames:
        wb.remove(wb[sheet])
    new_ws = wb.create_sheet(sheet, 0)
    for r in dataframe_to_rows(df, index=False, header=True):
        new_ws.append(r)
    wb.save(buf)
    buf.seek(0)
    st.session_state.download_buf = buf

# — Bouton de téléchargement
if st.session_state.completed:
    st.download_button(
        "💾 Télécharger Excel",
        data=st.session_state.download_buf,
        file_name="output.xlsx",
        mime="application/vnd.openxmlformats-officedocument-spreadsheetml.sheet"
    )

# — Journal de traitement
if st.session_state.log_entries:
    st.markdown("### 📑 Journal de traitement")
    log_df = pd.DataFrame(st.session_state.log_entries)
    st.dataframe(log_df)
    csv_data = log_df.to_csv(index=False).encode("utf-8")
    st.download_button("📝 Télécharger le journal (CSV)", data=csv_data,
                       file_name="journal.csv", mime="text/csv")
